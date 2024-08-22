from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.admin.views.decorators import staff_member_required
from orders.models import Order
from .models import DailySalesReport
from django.utils import timezone
from django.utils.timezone import now, make_aware, localtime
from datetime import timedelta, datetime
from decimal import Decimal
from django.core.serializers.json import DjangoJSONEncoder
from django.urls import reverse
from django.http import FileResponse, HttpResponseRedirect
from django.urls import reverse_lazy
import json
import calendar
import openpyxl
from openpyxl.utils import get_column_letter
from telegram_bot.bot import send_sales_report_via_telegram
import asyncio
import os
from django.conf import settings



@staff_member_required
def report_dashboard(request):
    return render(request, 'analytics/report_dashboard.html')

@staff_member_required
def daily_sales_report(request):
    try:
        current_time_utc = timezone.now() # Время в UTC
        current_time_local = localtime(current_time_utc) # Конвертация в локальное время
        current_date_local = current_time_local.date()
        today = current_date_local
        orders_today = Order.objects.filter(order_date__date=today)
        total_sales = sum(order.total_cost for order in orders_today)

        # Преобразование данных заказа в JSON-формат
        order_data = [
            {
                'order_id': order.id,
                'order_date': localtime(order.order_date).strftime('%Y-%m-%d %H:%M:%S'),
                'products': order.products,
                'total_cost': float(order.total_cost),
            }
            for order in orders_today
        ]
        order_data_json = json.dumps(order_data, cls=DjangoJSONEncoder)

        report, created = DailySalesReport.objects.get_or_create(
            report_date=today,
            defaults={'order_data': order_data_json, 'expenses': Decimal('0.00')}
        )

        if request.method == 'POST':
            expenses = request.POST.get('expenses')
            if expenses:
                report.expenses = Decimal(expenses)
                report.save()

            if 'send_report' in request.POST:
                file_name = f'sales_report_{today}.xlsx'
                file_path = os.path.join(settings.BASE_DIR, file_name)
                if os.path.exists(file_path):
                    # Вычисление значений total_expenses и profit
                    total_expenses = report.expenses
                    profit = total_sales - total_expenses
                    # Формируем подпись для отчета
                    caption = (
                        f"Отчёт о продажах за { today.strftime('%d.%m.%Y')}:\n"
                        f"Общая сумма продаж: {total_sales} руб.\n"
                        f"Общая сумма расходов: {total_expenses} руб.\n"
                        f"Прибыль: {profit} руб."
                    )

                    # Вызов функции отправки отчета
                    asyncio.run(send_sales_report_via_telegram(file_path, caption))

                return redirect('analytics:daily_sales_report')

        profit = total_sales - report.expenses

        # Генерация Excel-файла
        file_path = f'sales_report_{today}.xlsx'
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = f"Отчёт о продажах {today}"

        # Заполнение заголовков
        sheet["A1"] = f"Отчёт о продажах за {today}"
        sheet["A2"] = f"Общая сумма продаж: {total_sales} руб."
        sheet["A3"] = f"Общая сумма расходов: {report.expenses} руб."
        sheet["A4"] = f"Прибыль: {profit} руб."

        # Заголовок таблицы
        headers = ["Заказ", "Общая сумма", "Время", "Название товара", "Количество", "Цена"]
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            sheet[f"{col_letter}6"] = header

        # Заполнение данных заказа
        row_num = 7
        for order in order_data:
            for i, item in enumerate(order['products']):
                sheet[f"A{row_num}"] = order['order_id'] if i == 0 else ""
                sheet[f"B{row_num}"] = order['total_cost'] if i == 0 else ""
                sheet[f"C{row_num}"] = order['order_date'] if i == 0 else ""
                sheet[f"D{row_num}"] = item['product_name']
                sheet[f"E{row_num}"] = item['quantity']
                sheet[f"F{row_num}"] = item['price']
                row_num += 1

        # Сохранение файла на диск (перезаписываем, если файл уже существует)

        workbook.save(file_path)

        context = {
            'today': today,
            'orders': orders_today,
            'total_sales': total_sales,
            'total_expenses': report.expenses,
            'profit': profit,
            'report': report,
        }

    except Exception as e:
        messages.error(request, f"Произошла ошибка: {str(e)}")
        return redirect('analytics:report_dashboard')

    return render(request, 'analytics/daily_sales_report.html', context)

@staff_member_required
def monthly_sales_report(request):
    try:
        # Словарь для перевода названий месяцев
        month_translation = {
            'January': 'Январь',
            'February': 'Февраль',
            'March': 'Март',
            'April': 'Апрель',
            'May': 'Май',
            'June': 'Июнь',
            'July': 'Июль',
            'August': 'Август',
            'September': 'Сентябрь',
            'October': 'Октябрь',
            'November': 'Ноябрь',
            'December': 'Декабрь'
        }

        # Получаем текущую дату
        current_time_utc = timezone.now()  # Время в UTC
        current_time_local = localtime(current_time_utc)  # Конвертация в локальное время
        current_date_local = current_time_local.date()
        today = current_date_local

        # Получаем начальную и конечную даты из параметров запроса или устанавливаем значения по умолчанию
        start_date_str = request.GET.get('start_date')
        end_date_str = request.GET.get('end_date')

        if not start_date_str:
            start_date_str = today.replace(day=1).strftime('%d.%m.%Y')
        if not end_date_str:
            end_date_str = today.strftime('%d.%m.%Y')

        start_date = make_aware(datetime.strptime(start_date_str, '%d.%m.%Y'))
        end_date = make_aware(datetime.strptime(end_date_str, '%d.%m.%Y'))

        first_day_of_month = start_date.replace(day=1)
        last_day_of_month = first_day_of_month.replace(day=calendar.monthrange(start_date.year, start_date.month)[1])

        is_full_month = start_date.date() == first_day_of_month.date() and end_date.date() == last_day_of_month.date()

        # Получаем английское название месяца и переводим его на русский
        english_month_name = start_date.strftime('%B')
        month_name = month_translation.get(english_month_name, english_month_name)

        # Получаем заказы за указанный период
        orders = Order.objects.filter(order_date__range=(start_date, end_date + timedelta(days=1) - timedelta(seconds=1)))

        # Подсчитываем общие продажи
        total_sales = sum(order.total_cost for order in orders)

        # Подсчитываем общие расходы за указанный период, используя модель DailySalesReport
        daily_reports = DailySalesReport.objects.filter(report_date__range=(start_date, end_date))
        total_expenses = sum(report.expenses for report in daily_reports)

        # Подсчитываем прибыль
        profit = total_sales - total_expenses

        # Передаем данные в шаблон
        context = {
            'orders': orders,
            'total_sales': total_sales,
            'total_expenses': total_expenses,
            'profit': profit,
            'start_date': start_date_str,
            'end_date': end_date_str,
            'is_full_month': is_full_month,
            'month_name': month_name,  # Русское название месяца
        }
    except ValueError:
        messages.error(request, "Неверный формат даты. Используйте формат ДД.ММ.ГГГГ.")
        return redirect('analytics:monthly_sales_report')
    except Exception as e:
        messages.error(request, f"Произошла ошибка: {str(e)}")
        return redirect('analytics:report_dashboard')

    return render(request, 'analytics/monthly_sales_report.html', context)

@staff_member_required
def report_files_list(request):
    reports_directory = settings.BASE_DIR
    reports = []

    for file_name in os.listdir(reports_directory):
        if file_name.startswith('sales_report_') and file_name.endswith('.xlsx'):
            # Извлекаем дату из имени файла
            report_date_str = file_name.replace('sales_report_', '').replace('.xlsx', '')
            report_date = datetime.strptime(report_date_str, '%Y-%m-%d').date()
            formatted_date = report_date.strftime('%d.%m.%Y')

            reports.append({'file_name': file_name, 'report_date': formatted_date})

    context = {
        'reports': reports,
    }
    return render(request, 'analytics/report_files_list.html', context)

@staff_member_required
def download_report(request, file_name):
    file_path = os.path.join(settings.BASE_DIR, file_name)
    if os.path.exists(file_path):
        try:
            return FileResponse(open(file_path, 'rb'), as_attachment=True, filename=file_name)
        except IOError:
            messages.error(request, "Ошибка при чтении файла.")
    else:
        messages.error(request, "Файл не найден.")
    return redirect('analytics:report_files_list')

@staff_member_required
def delete_report(request, file_name):
    file_path = os.path.join(settings.BASE_DIR, file_name)
    if os.path.exists(file_path):
        try:
            os.remove(file_path)
            messages.success(request, "Файл успешно удален.")
        except OSError:
            messages.error(request, "Ошибка при удалении файла.")
    else:
        messages.error(request, "Файл не найден.")
    return HttpResponseRedirect(reverse_lazy('analytics:report_files_list'))
